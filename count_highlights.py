"""
Highlight-counting script for the Line-By-Line coding pass.


No Claude API calls -- this reads what's already highlighted in each
transcript, then:
  1. Writes one row per highlighted segment into the "Analysis" tab of
     Thematic Analysis Sheet (in the Customer Discovery folder, one level
     up from the Line-By-Line transcripts folder).
  2. Writes the total appearance count for each specific code into the
     "Frequency" tab of that same sheet.
  3. Still writes the local line_by_line_highlight_counts.md/.xlsx reports,
     rolled up by research group (Ethicality, Feasibility, Viability,
     Desirability).


WHY CATEGORY-LEVEL COUNTING IS THE RELIABLE PART:
In Codebook.xlsx, every code's highlight color is inherited from its row's
fill in column A, and that fill is shared across ALL codes in the same
research group (e.g. Ethics, Wellbeing, Inequality, Agency & Empowerment...
all share the same green). That means a highlighted quote's color tells you
its CATEGORY unambiguously, but not which of the ~10 codes in that category
it is -- that distinction only lived in the human coder's judgment when they
applied the highlight, not in the color itself. So:


  - Category counts (Theme column, category summary) are exact, derived
    straight from color.
  - Code-level counts (Segment ID, Frequency tab) are a best-effort bonus:
    if the highlighted text happens to contain exactly one codebook keyword
    from that category, we attribute it to that code. If it contains
    several, or none, the Segment ID/Initial Code are left blank/marked
    "(unspecified)" in the Analysis row, and it's NOT added to the
    Frequency tab (which is keyed by a single code) -- rather than
    guessing.


IDEMPOTENT RE-RUNS: before appending, the script reads what's already in
the Analysis tab and skips any (Interview ID, Quote) pair already present,
so running this twice won't duplicate rows.


Setup:
    pip install openpyxl google-api-python-client google-auth-oauthlib


    If you've run an older version of this script before, delete
    token.json first -- this version needs write access to Sheets, which
    your old token won't have been granted.


Usage:
    python count_highlights.py                  # whole Line-By-Line folder
    python count_highlights.py --doc-id XXXXX    # a single doc
    python count_highlights.py --dry-run         # skip writing to the Sheet
"""


import argparse
import math
import os
import re


import openpyxl
from googleapiclient.discovery import build


CODEBOOK_PATH = "Codebook.xlsx"


# From: https://drive.google.com/drive/folders/1nKCfNC8IOD9l435CfoBhE7nNpjveBfY_
LINE_BY_LINE_FOLDER_ID = "1nKCfNC8IOD9l435CfoBhE7nNpjveBfY_"


# The Thematic Analysis Sheet lives in the "Customer Discovery" folder,
# found by name rather than by ID/path since Drive search doesn't care how
# deep it's nested.
THEMATIC_SHEET_NAME = "Thematic Analysis Sheet"
ANALYSIS_TAB = "Analysis"
ANALYSIS_HEADER_ROW = 3  # header row inside the Analysis tab (data starts row 4)
FREQUENCY_TAB = "Frequency"
FREQUENCY_HEADER_ROW = 2  # header row inside the Frequency tab (data starts row 3)


SCOPES = ["https://www.googleapis.com/auth/documents.readonly",
          "https://www.googleapis.com/auth/drive.readonly",
          "https://www.googleapis.com/auth/spreadsheets"]


OUTPUT_MD = "line_by_line_highlight_counts.md"
OUTPUT_XLSX = "line_by_line_highlight_counts.xlsx"




# ---------------------------------------------------------------------------
# Codebook loading: codes + category-by-color mapping
# ---------------------------------------------------------------------------


def _rgb(hex6):
    h = hex6[-6:]
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))




def _dist(c1, c2):
    return math.sqrt(sum((a - b) ** 2 for a, b in zip(c1, c2)))




def load_codebook(path=CODEBOOK_PATH):
    """
    Reads Codebook.xlsx:
      - Legend at E2:E5 ("Ethicality", "Feasability", "Viability",
        "Desirablity"), color = fill of each legend cell.
      - Codes from row 9 down (Code ID | Code Name | Operational Definition
        | Included when... | Excluded when... | Example Quote), color =
        fill of the Code ID cell.


    Individual codes' fill colors are matched to the closest legend color
    via a greedy nearest-color assignment (exact matches first), grouping
    codes into their research category. Any color block with fewer than 2
    codes (e.g. a one-off like "Pain Point") is left as "Uncategorized"
    rather than force-matched, since a single row isn't a reliable cluster.


    Returns:
        codebook: {code_id: {"code_name", "color", "category", ...}}
        categories: [name, ...] in legend order
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]


    legend = []
    for r in range(2, 6):
        cell = ws.cell(row=r, column=5)
        fill = cell.fill.fgColor.rgb if cell.fill else None
        if cell.value and fill and fill != "00000000":
            legend.append((str(cell.value).strip(), _rgb(fill)))


    # Pull raw code rows.
    rows = []
    for row in ws.iter_rows(min_row=9, max_row=300):
        code_id_cell, name_cell, def_cell, inc_cell, exc_cell, ex_cell = row[:6]
        if code_id_cell.value is None or name_cell.value is None:
            continue
        fill = code_id_cell.fill.fgColor.rgb if code_id_cell.fill else None
        if not fill or fill == "00000000":
            continue
        rows.append({
            "code_id": str(code_id_cell.value).strip(),
            "code_name": str(name_cell.value).strip(),
            "color": f"#{fill[2:]}".lower(),
            "operational_definition": str(def_cell.value).strip() if def_cell.value else "",
            "included_when": str(inc_cell.value).strip() if inc_cell.value else "",
            "excluded_when": str(exc_cell.value).strip() if exc_cell.value else "",
            "example_quote": str(ex_cell.value).strip() if ex_cell.value else "",
        })


    # Group rows by their raw fill color, in order of first appearance.
    blocks = {}
    block_order = []
    for r in rows:
        if r["color"] not in blocks:
            blocks[r["color"]] = []
            block_order.append(r["color"])
        blocks[r["color"]].append(r)


    # Greedy nearest-color assignment: exact/closest matches win first,
    # each legend category and each color block used at most once.
    candidates = []
    for color in block_order:
        if len(blocks[color]) < 2:
            continue  # singleton block, don't force it into a category
        for cat_name, cat_rgb in legend:
            candidates.append((_dist(_rgb(color), cat_rgb), color, cat_name))
    candidates.sort(key=lambda x: x[0])


    color_to_category = {}
    used_categories = set()
    for _, color, cat_name in candidates:
        if color in color_to_category or cat_name in used_categories:
            continue
        color_to_category[color] = cat_name
        used_categories.add(cat_name)


    codebook = {}
    for r in rows:
        r["category"] = color_to_category.get(r["color"], "Uncategorized")
        codebook[r["code_id"]] = r


    categories = [name for name, _ in legend] + (
        ["Uncategorized"] if any(v["category"] == "Uncategorized" for v in codebook.values()) else []
    )
    return codebook, categories




def color_to_codes(codebook, hex_color):
    """All codes sharing this highlight color (usually a whole category)."""
    hex_color = hex_color.lower()
    return [c for c in codebook.values() if c["color"] == hex_color]




def build_keyword_pattern(keyword):
    """Same lemma-ish matching as the tagging script, used only to try to
    narrow a highlighted quote down to one specific code within its color
    group -- never to invent a match that wasn't actually highlighted."""
    parts = keyword.split()
    last = parts[-1].lower()
    for suf in ("ing", "edly", "ed", "es", "s"):
        if last.endswith(suf) and len(last) - len(suf) >= 3:
            last = last[: -len(suf)]
            break
    variant = re.escape(last) + r"(?:s|es|d|ed|ing)?"
    if len(parts) > 1:
        prefix = r"\s+".join(re.escape(p) for p in parts[:-1])
        return r"\b" + prefix + r"\s+" + variant + r"\b"
    return r"\b" + variant + r"\b"




def guess_specific_code(quote, same_color_codes):
    """If exactly one of the codes sharing this highlight's color has its
    keyword literally present in the quoted text, attribute the highlight
    to that code. Otherwise leave it unspecified -- guessing among several
    equally-plausible codes would be fabricating precision we don't have."""
    hits = []
    for c in same_color_codes:
        pattern = build_keyword_pattern(c["code_name"])
        if re.search(pattern, quote, flags=re.IGNORECASE):
            hits.append(c)
    if len(hits) == 1:
        return hits[0]["code_id"]
    return None




# ---------------------------------------------------------------------------
# Google auth / Drive folder listing / doc structure reading
# (unchanged approach from the tagging script; read-only scopes here)
# ---------------------------------------------------------------------------


def get_credentials(credentials_path="credentials.json", token_path="token.json"):
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow


    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as f:
            f.write(creds.to_json())
    return creds




def get_docs_in_folder(folder_id, creds):
    """
    Lists every item directly inside the given folder (no subfolders --
    this folder is flat). Returns Google Docs to process, plus a list of
    anything else found (e.g. uploaded .docx files, Sheets) that this
    script can't read, so they can be reported rather than silently missed.


    Returns: (docs, skipped)
        docs:    [{"id", "name"}, ...]  -- native Google Docs
        skipped: [{"name", "mimeType"}, ...]  -- everything else
    """
    service = build("drive", "v3", credentials=creds)
    query = f"'{folder_id}' in parents and trashed=false"
    items, page_token = [], None
    while True:
        results = service.files().list(
            q=query, spaces="drive", fields="nextPageToken, files(id, name, mimeType)",
            pageToken=page_token, supportsAllDrives=True,
            includeItemsFromAllDrives=True, corpora="allDrives",
        ).execute()
        items.extend(results.get("files", []))
        page_token = results.get("nextPageToken")
        if not page_token:
            break


    docs, skipped = [], []
    for item in items:
        if item["mimeType"] == "application/vnd.google-apps.document":
            docs.append({"id": item["id"], "name": item["name"]})
        else:
            skipped.append({"name": item["name"], "mimeType": item["mimeType"]})
    return docs, skipped




def find_spreadsheet_by_name(name, creds):
    """Finds a Google Sheet by exact name anywhere the credentialed account
    can see (no need to know its folder path). Returns the spreadsheet ID,
    or None if not found. If multiple matches exist, uses the most recently
    modified one and warns."""
    service = build("drive", "v3", credentials=creds)
    query = (f"name='{name}' and mimeType='application/vnd.google-apps.spreadsheet' "
             f"and trashed=false")
    results = service.files().list(
        q=query, spaces="drive", fields="files(id, name, modifiedTime)",
        orderBy="modifiedTime desc", supportsAllDrives=True,
        includeItemsFromAllDrives=True, corpora="allDrives",
    ).execute()
    files = results.get("files", [])
    if not files:
        return None
    if len(files) > 1:
        print(f"  Warning: found {len(files)} files named '{name}', using the most recently modified.")
    return files[0]["id"]




def extract_interview_id(doc_name):
    """Pulls the leading interview number off a transcript name, e.g.
    '72 - Cheng - Malwa, Mildred' -> 72. Falls back to the full name if no
    leading number is found."""
    m = re.match(r"\s*(\d+)", doc_name)
    return int(m.group(1)) if m else doc_name




def fetch_doc(doc_id, creds):
    service = build("docs", "v1", credentials=creds)
    return service.documents().get(documentId=doc_id).execute()




def walk_doc_runs(doc):
    """Flattens the doc into text runs with real startIndex/endIndex and any
    existing highlight background color -- see tagging script for details."""
    runs = []
    for element in doc.get("body", {}).get("content", []):
        paragraph = element.get("paragraph")
        if not paragraph:
            continue
        for run in paragraph.get("elements", []):
            text_run = run.get("textRun")
            if not text_run:
                continue
            content = text_run.get("content", "")
            if not content:
                continue
            style = text_run.get("textStyle", {})
            bg = style.get("backgroundColor", {}).get("color", {}).get("rgbColor")
            bg_hex = None
            if bg:
                r = round(bg.get("red", 0) * 255)
                g = round(bg.get("green", 0) * 255)
                b = round(bg.get("blue", 0) * 255)
                bg_hex = f"#{r:02x}{g:02x}{b:02x}"
            runs.append({"text": content, "start": run.get("startIndex"),
                         "end": run.get("endIndex"), "bg_hex": bg_hex})
    return runs




def find_existing_highlighted_spans(runs, codebook):
    """Merges consecutive same-color runs into spans and attaches every code
    that shares that highlight color."""
    known_colors = {v["color"] for v in codebook.values()}
    spans, current = [], None
    for run in runs:
        color = run["bg_hex"].lower() if run["bg_hex"] else None
        if color in known_colors:
            if current and current["color"] == color and current["end"] == run["start"]:
                current["text"] += run["text"]
                current["end"] = run["end"]
            else:
                if current:
                    spans.append(current)
                current = {"text": run["text"], "start": run["start"], "end": run["end"], "color": color}
        else:
            if current:
                spans.append(current)
                current = None
    if current:
        spans.append(current)
    for span in spans:
        span["codes"] = color_to_codes(codebook, span["color"])
        span["category"] = span["codes"][0]["category"] if span["codes"] else "Unknown color"
    return [s for s in spans if s["text"].strip()]




# ---------------------------------------------------------------------------
# Per-doc counting
# ---------------------------------------------------------------------------


def count_doc(doc_id, doc_name, creds, codebook):
    doc = fetch_doc(doc_id, creds)
    runs = walk_doc_runs(doc)
    spans = find_existing_highlighted_spans(runs, codebook)


    category_counts = {}
    code_counts = {}  # code_id -> count ("UNSPECIFIED::<category>" bucket for ambiguous ones)
    analysis_rows = []  # one row per highlighted segment, for the Analysis tab
    interview_id = extract_interview_id(doc_name)


    for span in spans:
        category_counts[span["category"]] = category_counts.get(span["category"], 0) + 1


        specific = guess_specific_code(span["text"], span["codes"]) if span["codes"] else None
        key = specific if specific else f"UNSPECIFIED::{span['category']}"
        code_counts[key] = code_counts.get(key, 0) + 1


        quote = span["text"].strip()
        if specific:
            segment_id, initial_code = specific, codebook[specific]["code_name"]
        else:
            segment_id, initial_code = "", f"(unspecified -- {span['category']} color only)"
        analysis_rows.append({
            "interview_id": interview_id,
            "segment_id": segment_id,
            "quote": quote,
            "initial_code": initial_code,
            "theme": span["category"],
        })


    return {"doc_name": doc_name, "total": len(spans),
            "category_counts": category_counts, "code_counts": code_counts,
            "analysis_rows": analysis_rows}




# ---------------------------------------------------------------------------
# Writing to the Thematic Analysis Sheet (Analysis + Frequency tabs)
# ---------------------------------------------------------------------------


def get_existing_analysis_keys(spreadsheet_id, creds):
    """Reads (Interview ID, Quote/Paraphrase) pairs already present in the
    Analysis tab, so re-running the script doesn't append duplicate rows."""
    service = build("sheets", "v4", credentials=creds)
    rng = f"{ANALYSIS_TAB}!A{ANALYSIS_HEADER_ROW + 1}:E"
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range=rng).execute()
    keys = set()
    for row in result.get("values", []):
        interview_id = row[0] if len(row) > 0 else ""
        quote = row[4] if len(row) > 4 else ""
        if interview_id or quote:
            keys.add((str(interview_id).strip(), str(quote).strip()))
    return keys




def append_analysis_rows(spreadsheet_id, creds, rows):
    """Appends new rows to the Analysis tab: Interview ID | Segment ID |
    Goal/Objective | Participant Type | Quote/Paraphrase | Initial Code |
    Theme | Observation. Goal/Objective, Participant Type, and Observation
    are left blank for a human to fill in."""
    if not rows:
        return 0
    service = build("sheets", "v4", credentials=creds)
    values = [
        [r["interview_id"], r["segment_id"], "", "", r["quote"], r["initial_code"], r["theme"], ""]
        for r in rows
    ]
    service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=f"{ANALYSIS_TAB}!A{ANALYSIS_HEADER_ROW}:H",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={"values": values},
    ).execute()
    return len(values)




def write_frequency_counts(spreadsheet_id, creds, grand_code):
    """Writes total appearance counts into the Frequency tab's 'Number of
    Apperances' column, matched by Code ID -- reads the sheet's own Code ID
    column first so it works regardless of row order, and only counts
    highlights that were confidently attributed to one specific code
    (UNSPECIFIED:: buckets are excluded, since Frequency is per-code)."""
    service = build("sheets", "v4", credentials=creds)
    id_range = f"{FREQUENCY_TAB}!A{FREQUENCY_HEADER_ROW + 1}:A"
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range=id_range).execute()
    code_ids = [row[0].strip() if row else "" for row in result.get("values", [])]
    if not code_ids:
        print("  Warning: Frequency tab has no Code ID rows -- skipping frequency write.")
        return 0


    counts = [[grand_code.get(cid, 0)] if cid else [""] for cid in code_ids]
    last_row = FREQUENCY_HEADER_ROW + len(code_ids)
    count_range = f"{FREQUENCY_TAB}!C{FREQUENCY_HEADER_ROW + 1}:C{last_row}"
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id, range=count_range,
        valueInputOption="RAW", body={"values": counts},
    ).execute()
    return sum(1 for c in counts if c[0])




# ---------------------------------------------------------------------------
# Report writers
# ---------------------------------------------------------------------------


def write_markdown_report(results, codebook, categories, skipped=None, path=OUTPUT_MD):
    skipped = skipped or []
    lines = ["# Line-By-Line Highlight Counts\n"]


    grand_category = {}
    grand_code = {}
    for res in results:
        for cat, n in res["category_counts"].items():
            grand_category[cat] = grand_category.get(cat, 0) + n
        for code_key, n in res["code_counts"].items():
            grand_code[code_key] = grand_code.get(code_key, 0) + n


    lines.append("## Overall totals by research group\n")
    grand_total = sum(grand_category.values())
    for cat in categories:
        n = grand_category.get(cat, 0)
        pct = f"{100 * n / grand_total:.1f}%" if grand_total else "0%"
        lines.append(f"- **{cat}**: {n} highlight(s) ({pct})")
    lines.append(f"- **Total highlights across all docs**: {grand_total}\n")


    lines.append("## Overall totals by code (within each group)\n")
    for cat in categories:
        cat_codes = {k: v for k, v in grand_code.items()
                     if (k.startswith("UNSPECIFIED::") and k.split("::", 1)[1] == cat)
                     or (k in codebook and codebook[k]["category"] == cat)}
        if not cat_codes:
            continue
        lines.append(f"\n**{cat}**")
        for k, v in sorted(cat_codes.items(), key=lambda x: -x[1]):
            if k.startswith("UNSPECIFIED::"):
                lines.append(f"- (unspecified code -- category color only): {v}")
            else:
                lines.append(f"- {codebook[k]['code_name']} ({k}): {v}")


    lines.append("\n## Per-document breakdown\n")
    for res in results:
        lines.append(f"### {res['doc_name']}")
        lines.append(f"Total highlights: {res['total']}")
        for cat in categories:
            n = res["category_counts"].get(cat, 0)
            if n:
                lines.append(f"- {cat}: {n}")
        lines.append("")


    if skipped:
        lines.append("## Skipped files (not native Google Docs, not read by this script)\n")
        for s in skipped:
            lines.append(f"- {s['name']} ({s['mimeType']})")
        lines.append("")


    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))




def write_xlsx_report(results, codebook, categories, skipped=None, path=OUTPUT_XLSX):
    from openpyxl.styles import Font


    skipped = skipped or []
    wb = openpyxl.Workbook()


    # --- Sheet 1: Category Summary (per document) ---
    ws1 = wb.active
    ws1.title = "Category Summary"
    ws1.append(["Category"] + [r["doc_name"] for r in results] + ["Total"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for cat in categories:
        row = [cat]
        total = 0
        for res in results:
            n = res["category_counts"].get(cat, 0)
            row.append(n)
            total += n
        row.append(total)
        ws1.append(row)
    ndocs = len(results)
    total_row = ["Total"]
    for col in range(2, ndocs + 3):
        col_letter = openpyxl.utils.get_column_letter(col)
        total_row.append(f"=SUM({col_letter}2:{col_letter}{1 + len(categories)})")
    ws1.append(total_row)
    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)


    # --- Sheet 2: Code-Level Detail ---
    ws2 = wb.create_sheet("Code Detail")
    ws2.append(["Category", "Code ID", "Code Name"] + [r["doc_name"] for r in results] + ["Total"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)


    for cat in categories:
        # named codes in this category
        cat_code_ids = [cid for cid, c in codebook.items() if c["category"] == cat]
        for cid in cat_code_ids:
            row = [cat, cid, codebook[cid]["code_name"]]
            total = 0
            for res in results:
                n = res["code_counts"].get(cid, 0)
                row.append(n)
                total += n
            row.append(total)
            if total:  # skip rows that never appear in any transcript
                ws2.append(row)
        # unspecified bucket for this category
        unspecified_total = sum(res["code_counts"].get(f"UNSPECIFIED::{cat}", 0) for res in results)
        if unspecified_total:
            row = [cat, "", "(unspecified -- category color only)"]
            total = 0
            for res in results:
                n = res["code_counts"].get(f"UNSPECIFIED::{cat}", 0)
                row.append(n)
                total += n
            row.append(total)
            ws2.append(row)


    # --- Sheet 3: Skipped Files ---
    if skipped:
        ws3 = wb.create_sheet("Skipped Files")
        ws3.append(["File name", "File type", "Reason"])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
        for s in skipped:
            ws3.append([s["name"], s["mimeType"], "Not a native Google Doc -- not read by this script"])


    sheets_to_size = [ws1, ws2] + ([ws3] if skipped else [])
    for ws in sheets_to_size:
        for col in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 45)


    wb.save(path)




# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Count Line-By-Line highlights by research group")
    parser.add_argument("--folder-id", default=LINE_BY_LINE_FOLDER_ID)
    parser.add_argument("--doc-id", default=None, help="Count a single doc instead of the whole folder")
    parser.add_argument("--codebook", default=CODEBOOK_PATH)
    parser.add_argument("--sheet-id", default=None,
                         help="Spreadsheet ID for Thematic Analysis Sheet (auto-found by name if omitted)")
    parser.add_argument("--dry-run", action="store_true",
                         help="Compute everything and write the local report, but don't touch the Google Sheet")
    args = parser.parse_args()


    codebook, categories = load_codebook(args.codebook)
    creds = get_credentials()


    if args.doc_id:
        docs, skipped = [{"id": args.doc_id, "name": args.doc_id}], []
    else:
        docs, skipped = get_docs_in_folder(args.folder_id, creds)
        print(f"Found {len(docs)} Google Doc(s) in the folder.")
        if skipped:
            print(f"Skipping {len(skipped)} file(s) that aren't native Google Docs "
                  f"(e.g. uploaded .docx, Sheets): " + ", ".join(s["name"] for s in skipped))


    results = []
    for d in docs:
        print(f"Counting highlights in {d['name']}...")
        try:
            results.append(count_doc(d["id"], d["name"], creds, codebook))
        except Exception as e:
            print(f"  -> Error: {e}")
            results.append({"doc_name": d["name"], "total": 0, "category_counts": {},
                             "code_counts": {}, "analysis_rows": []})


    write_markdown_report(results, codebook, categories, skipped=skipped)
    write_xlsx_report(results, codebook, categories, skipped=skipped)
    print(f"\nWrote {OUTPUT_MD} and {OUTPUT_XLSX}")


    if args.dry_run:
        print("\n--dry-run set: not writing to the Thematic Analysis Sheet.")
        return


    sheet_id = args.sheet_id or find_spreadsheet_by_name(THEMATIC_SHEET_NAME, creds)
    if not sheet_id:
        print(f"\nCouldn't find a Google Sheet named '{THEMATIC_SHEET_NAME}' -- "
              f"skipping the Analysis/Frequency write. Pass --sheet-id to point at it directly.")
        return


    print(f"\nWriting to Thematic Analysis Sheet ({sheet_id})...")


    all_rows = [r for res in results for r in res["analysis_rows"]]
    existing_keys = get_existing_analysis_keys(sheet_id, creds)
    new_rows = [r for r in all_rows
                if (str(r["interview_id"]).strip(), r["quote"].strip()) not in existing_keys]
    skipped_dupes = len(all_rows) - len(new_rows)


    n_appended = append_analysis_rows(sheet_id, creds, new_rows)
    print(f"  Analysis tab: appended {n_appended} new row(s)"
          + (f", skipped {skipped_dupes} already present" if skipped_dupes else "") + ".")


    grand_code = {}
    for res in results:
        for code_key, n in res["code_counts"].items():
            if not code_key.startswith("UNSPECIFIED::"):
                grand_code[code_key] = grand_code.get(code_key, 0) + n
    n_freq = write_frequency_counts(sheet_id, creds, grand_code)
    print(f"  Frequency tab: wrote counts for all codes; {n_freq} code(s) had at least one highlight.")




if __name__ == "__main__":
    main()
